import os
from openpyxl import load_workbook
from datetime import datetime
import zipfile

def cleanup():
    file = "schedules.xlsx"
    if not os.path.exists(file):
        return

    try:
        wb = load_workbook(file)
        today = datetime.today().date()

        for sheet in wb.sheetnames:
            try:
                sheet_date = datetime.strptime(sheet, "%Y-%m-%d").date()
                if sheet_date < today:
                    wb.remove(wb[sheet])
            except ValueError:
                continue  # Skip sheets without valid date names

        wb.save(file)

    except zipfile.BadZipFile:
        print("⚠️ Corrupted schedules.xlsx detected. Deleting it.")
        os.remove(file)
    except Exception as e:
        print(f"Unexpected error during cleanup: {e}")
