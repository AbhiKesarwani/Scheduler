import pandas as pd
import os
from openpyxl import load_workbook

SCHEDULE_FILE = "schedules.xlsx"

def write_schedule(new_entry):
    date_sheet = new_entry["Date"]
    new_df = pd.DataFrame([new_entry])

    if not os.path.exists(SCHEDULE_FILE):
        with pd.ExcelWriter(SCHEDULE_FILE, engine="openpyxl", mode="w") as writer:
            new_df.to_excel(writer, sheet_name=date_sheet, index=False)
    else:
        book = load_workbook(SCHEDULE_FILE)
        sheets = book.sheetnames

        # Read existing or add new
        if date_sheet in sheets:
            existing_df = pd.read_excel(SCHEDULE_FILE, sheet_name=date_sheet, engine="openpyxl")
            combined = pd.concat([existing_df, new_df], ignore_index=True)
        else:
            combined = new_df

        # Remove old version of sheet
        if date_sheet in sheets:
            book.remove(book[date_sheet])

        # Collect all sheets
        all_data = {sheet: pd.read_excel(SCHEDULE_FILE, sheet_name=sheet, engine="openpyxl")
                    for sheet in sheets if sheet != date_sheet}
        all_data[date_sheet] = combined

        sorted_sheets = sorted(all_data.keys())

        with pd.ExcelWriter(SCHEDULE_FILE, engine="openpyxl", mode="w") as writer:
            for sheet in sorted_sheets:
                all_data[sheet].to_excel(writer, sheet_name=sheet, index=False)

def read_schedule_sheet(date_sheet):
    try:
        return pd.read_excel(SCHEDULE_FILE, sheet_name=date_sheet, engine="openpyxl")
    except Exception:
        return pd.DataFrame()

def remove_schedule_entry(date_sheet, user, class_name, start_time):
    import pandas as pd
    from openpyxl import load_workbook

    try:
        df = pd.read_excel("schedules.xlsx", sheet_name=date_sheet, engine="openpyxl")
        original_len = len(df)

        # Make sure times are compared as strings
        filtered_df = df[~((df["User"] == user) & 
                           (df["Class"] == class_name) & 
                           (df["Start"] == start_time))]

        if len(filtered_df) == original_len:
            return False

        # Remove sheet and rewrite
        book = load_workbook("schedules.xlsx")
        if date_sheet in book.sheetnames:
            book.remove(book[date_sheet])
        with pd.ExcelWriter("schedules.xlsx", engine="openpyxl", mode="a") as writer:
            filtered_df.to_excel(writer, sheet_name=date_sheet, index=False)

        return True
    except Exception as e:
        print(f"Error deleting entry: {e}")
        return False

        return True
    except Exception as e:
        print(f"Error deleting: {e}")
        return False